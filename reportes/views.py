from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse
from django.utils.timezone import now
from datetime import timedelta

from carrito.models import Pedido
from registration.models import PerfilUsuario
from api.models import Producto
from vendedor.models import VentaFisica

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import DataBarRule


def get_or_create_perfil(user):
    perfil = getattr(user, 'perfilusuario', None)
    if not perfil:
        perfil, _ = PerfilUsuario.objects.get_or_create(user=user)
    return perfil


def normalizar_venta(venta, tipo):
    """ Unifica ventas físicas y Webpay en un mismo formato """
    if tipo == "webpay":
        return {
            "id": venta.id,
            "fecha": venta.fecha,
            "cliente": venta.cliente,
            "total": float(venta.total),
            "estado": venta.estado,
            "productos": [p.nombre for p in venta.productos.all()],
        }

    if tipo == "fisica":
        return {
            "id": venta.id,
            "fecha": venta.fecha,
            "cliente": venta.cliente or "Cliente presencial",
            "total": float(venta.total),
            "estado": f"Venta física – {venta.metodo_pago}",
            "productos": [item.producto.nombre for item in venta.items.all()],
        }


class ReporteVentasDiariasAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        perfil = get_or_create_perfil(request.user)
        if perfil.rol.nombre != "Administrador":
            return Response({"error": "No autorizado"}, status=403)

        hoy = now().date()
        inicio = hoy
        fin = hoy + timedelta(days=1)

        webpay = Pedido.objects.filter(
            estado="Pagado",
            fecha__range=[inicio, fin]
        )

        fisicas = VentaFisica.objects.filter(
            fecha__range=[inicio, fin]
        )

        data = (
            [normalizar_venta(v, "webpay") for v in webpay] +
            [normalizar_venta(v, "fisica") for v in fisicas]
        )

        return Response(data)


class ReporteVentasSemanalesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        perfil = get_or_create_perfil(request.user)
        if perfil.rol.nombre != "Administrador":
            return Response({"error": "No autorizado"}, status=403)

        hoy = now()
        inicio = hoy - timedelta(days=7)

        webpay = Pedido.objects.filter(
            estado="Pagado",
            fecha__gte=inicio
        )

        fisicas = VentaFisica.objects.filter(
            fecha__gte=inicio
        )

        data = (
            [normalizar_venta(v, "webpay") for v in webpay] +
            [normalizar_venta(v, "fisica") for v in fisicas]
        )

        return Response(data)


class ReporteVentasMensualesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        perfil = get_or_create_perfil(request.user)
        if perfil.rol.nombre != "Administrador":
            return Response({"error": "No autorizado"}, status=403)

        hoy = now()
        inicio_mes = hoy.replace(day=1, hour=0, minute=0)

        webpay = Pedido.objects.filter(
            estado="Pagado",
            fecha__gte=inicio_mes
        )

        fisicas = VentaFisica.objects.filter(
            fecha__gte=inicio_mes
        )

        data = (
            [normalizar_venta(v, "webpay") for v in webpay] +
            [normalizar_venta(v, "fisica") for v in fisicas]
        )

        return Response(data)


class ReporteExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        perfil = get_or_create_perfil(request.user)
        if perfil.rol.nombre != "Administrador":
            return Response({"error": "No autorizado"}, status=403)

        tipo = request.GET.get("tipo", "mensual")
        hoy = now()

        if tipo == "diario":
            inicio = hoy.date()
            fin = inicio + timedelta(days=1)
        elif tipo == "semanal":
            inicio = hoy - timedelta(days=7)
            fin = hoy
        else:
            inicio = hoy.replace(day=1, hour=0, minute=0)
            fin = hoy

        webpay = Pedido.objects.filter(
            estado="Pagado",
            fecha__range=[inicio, fin]
        )

        fisicas = VentaFisica.objects.filter(
            fecha__range=[inicio, fin]
        )

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Ventas"

        ws1.append(["ID", "Fecha", "Cliente", "Productos", "Total", "Estado"])
        for c in ws1[1]:
            c.font = Font(bold=True)

        total_productos = 0

        ventas = (
            [normalizar_venta(v, "webpay") for v in webpay] +
            [normalizar_venta(v, "fisica") for v in fisicas]
        )

        for v in ventas:
            ws1.append([
                v["id"],
                v["fecha"].strftime("%Y-%m-%d %H:%M"),
                v["cliente"],
                ", ".join(v["productos"]),
                float(v["total"]),
                v["estado"]
            ])

            total_productos += len(v["productos"])

        ws2 = wb.create_sheet("Resumen")
        ws2["A1"] = f"Resumen del período ({tipo})"
        ws2["A1"].font = Font(size=14, bold=True)

        ws2.append([""])
        ws2.append(["Total de ventas", len(ventas)])
        ws2.append(["Total productos vendidos", total_productos])
        ws2.append(["Total dinero recaudado", sum(v["total"] for v in ventas)])

        ws3 = wb.create_sheet("Reposición")
        ws3.append(["Producto", "Stock", "Estado", "Prioridad"])

        for c in ws3[1]:
            c.font = Font(bold=True)

        row = 2
        productos = Producto.objects.all()

        for p in productos:
            if p.stock == 0:
                estado = "REPOSICIÓN URGENTE - AGOTADO"
                color = "FF0000"
                prioridad = 100

            elif 1 <= p.stock <= 5:
                estado = "Producto por agotarse"
                color = "FFA500"
                prioridad = 70

            else:
                estado = "OK"
                color = "00CC00"
                prioridad = 10

            ws3.append([p.nombre, p.stock, estado, prioridad])

            ws3[f"C{row}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

            row += 1

        ws3.conditional_formatting.add(
            f"D2:D{row}",
            DataBarRule(start_type="num", end_type="num", start_value=0, end_value=100, color="638EC6")
        )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="reporte_{tipo}.xlsx"'

        wb.save(response)
        return response
